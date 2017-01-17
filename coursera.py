from bs4 import BeautifulSoup
from openpyxl import Workbook
from lxml import etree
import requests
import argparse
import random
import json


COURSES_URL = "https://www.coursera.org/sitemap~www~courses.xml"


def get_path_argument():
    parser = argparse.ArgumentParser(description='Output file for information about <Coursera> courses')
    parser.add_argument('file', type=str, help='specify .xlsx file to output')
    return parser.parse_args().file


def fetch_url_content(url):
    return requests.get(url).content


def get_courses_random_sample(courses_raw_info, sample_size=20):
    root = etree.fromstring(courses_raw_info)
    return [course_unit[0].text for course_unit in random.sample(list(root), sample_size)]


def fetch_course_start_date(beautiful_soup_object):
    start_date_footprint = 'startDate'
    json_with_possible_start_date = beautiful_soup_object.find('script', {'type': 'application/ld+json'})
    if json_with_possible_start_date and start_date_footprint in json_with_possible_start_date.text:
        datetime = json.loads(json_with_possible_start_date.text)['hasCourseInstance'][0][start_date_footprint]
        return datetime


def fetch_course_name(beautiful_soup_object):
    course_name = beautiful_soup_object.find('div', {'class': 'title display-3-text'})
    if course_name:
        return course_name.text


def fetch_course_rate(beautiful_soup_object):
    course_rate = beautiful_soup_object.find('div', {'class': 'ratings-text bt3-visible-xs'})
    return course_rate.text if course_rate else None


def fetch_course_language(beautiful_soup_object):
    course_language = beautiful_soup_object.find('div', {'class': 'language-info'})
    if course_language:
        return course_language.text


def get_course_info(course_page):
        soup_object = BeautifulSoup(course_page, 'html.parser')
        duration = soup_object.find_all('div', {'class': 'week'})
        course_name = fetch_course_name(soup_object)
        course_rate = fetch_course_rate(soup_object)
        course_language = fetch_course_language(soup_object)
        course_duration = len(duration) if duration else None
        course_start_date = fetch_course_start_date(soup_object)
        course_url = soup_object.find('meta', property='og:url')['content']

        return course_name, course_url, course_language, course_start_date, course_duration, course_rate


def output_courses_info_to_xlsx(courses_list, filepath):
    courses_workbook = Workbook()
    workbook_page = courses_workbook.active
    column_names = ['Course name', 'URL', 'Language', 'Start date', 'Duration', 'Average rating']

    for index, name in enumerate(column_names, 1):
        workbook_page.cell(row=1, column=index).value = name

    for position, course_info in enumerate(courses_list, 2):
        for fetched_parameter_id, fetched_parameter in enumerate(course_info, 1):
            workbook_page.cell(row=position, column=fetched_parameter_id).value = fetched_parameter if \
                fetched_parameter else 'No info'

    courses_workbook.save(filepath)
    return True


if __name__ == '__main__':
    print('Collecting information about courses... wait a minute please...')

    courses_links = get_courses_random_sample(fetch_url_content(COURSES_URL))
    fetched_pages = [fetch_url_content(url) for url in courses_links]
    parsed_courses_info = [get_course_info(page) for page in fetched_pages]

    if output_courses_info_to_xlsx(parsed_courses_info, get_path_argument()):
        print('File saved to {}'.format(get_path_argument()))
